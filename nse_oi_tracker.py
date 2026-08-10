#!/usr/bin/env python3
"""
NSE OI Tracker
==============
Har 5 minute par NSE se F&O stocks ka Open Interest data uthata hai (9:15 AM se 3:30 PM),
har snapshot ko ek alag Excel sheet me daalta hai, aur din khatam hone par ek
"CONCLUSION" sheet banata hai:

    Symbol | 09.15 | 09.20 | 09.25 | ... | 15.30
    TATASTEEL | 2.04 | 8.90 | 5.90 | ... | 3.11

Values = % change in OI us timestamp par.

Usage
-----
    python nse_oi_tracker.py                 # aaj ka session chalao (9:15 -> 3:30)
    python nse_oi_tracker.py --daemon        # hamesha chalta rahe, roz apne aap
    python nse_oi_tracker.py --once          # abhi ek snapshot lo (testing ke liye)
    python nse_oi_tracker.py --rebuild FILE  # purani file se CONCLUSION dubara banao
    python nse_oi_tracker.py --mock          # bina internet ke test (fake data)
"""

import argparse
import json
import logging
import os
import random
import sys
import time
from datetime import date, datetime, timedelta
from datetime import time as dtime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo
except ImportError:  # Python 3.8
    from backports.zoneinfo import ZoneInfo

# Server UTC me ho sakta hai — IST force karna zaroori hai
IST = ZoneInfo("Asia/Kolkata")


def now_ist():
    return datetime.now(IST)


def today_ist():
    return now_ist().date()


def at_ist(d, t):
    return datetime.combine(d, t, tzinfo=IST)

# ------------------------------------------------------------------ CONFIG --

OUTPUT_DIR = os.getenv(
    "OI_OUTPUT_DIR",
    os.path.join(os.path.expanduser("~"), "NSE_OI_Data"),
)

MARKET_OPEN = dtime(9, 15)
MARKET_CLOSE = dtime(15, 30)
INTERVAL_MIN = 5

# "day"      -> % change in OI vs previous day's close OI (NSE ka apna number)
# "interval" -> % change in OI vs pichhle 5-minute snapshot (intraday momentum)
CHANGE_MODE = "day"

# ---- Cloud upload (S3) ----
# Credentials environment se aayenge — script me kabhi mat likhna.
#   export AWS_ACCESS_KEY_ID=...
#   export AWS_SECRET_ACCESS_KEY=...
UPLOAD_TO_S3 = os.getenv("OI_S3_BUCKET") is not None
S3_BUCKET = os.getenv("OI_S3_BUCKET", "")
S3_PREFIX = os.getenv("OI_S3_PREFIX", "nse-oi-data")
S3_REGION = os.getenv("OI_S3_REGION", "ap-south-1")

# Snapshot sheet me kaunse columns save karne hain
SNAPSHOT_COLS = [
    "symbol", "baselineOI", "latestOI", "changeInOI", "pctChangeOI",
    "volume", "valueInCr", "underlyValue",
]

# NSE market holidays (har saal update kar lena)
HOLIDAYS_2026 = {
    "2026-01-26", "2026-02-15", "2026-03-03", "2026-03-04", "2026-03-21",
    "2026-04-01", "2026-04-03", "2026-04-14", "2026-05-01", "2026-05-27",
    "2026-08-15", "2026-09-14", "2026-10-02", "2026-10-20", "2026-11-09",
    "2026-11-10", "2026-11-24", "2026-12-25",
}

BASE = "https://www.nseindia.com"
OI_URL = f"{BASE}/api/live-analysis-oi-spurts-underlyings"
REFERER = f"{BASE}/market-data/oi-spurts"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": REFERER,
    "Connection": "keep-alive",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
}

# ----------------------------------------------------------------- LOGGING --


def setup_logging():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    log_path = os.path.join(OUTPUT_DIR, "oi_tracker.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return log_path


log = logging.getLogger(__name__)

# -------------------------------------------------------------- NSE CLIENT --


class NSEClient:
    """NSE cookies maangta hai, isliye pehle homepage hit karke session banate hain."""

    def __init__(self, mock=False):
        self.mock = mock
        self.session = None
        if not mock:
            self._new_session()

    def _new_session(self):
        s = requests.Session()
        s.headers.update(HEADERS)
        # Cookie warm-up: homepage -> OI spurts page
        s.get(BASE, timeout=15)
        time.sleep(1)
        s.get(REFERER, timeout=15)
        time.sleep(1)
        self.session = s
        log.info("NSE session ready (cookies mil gaye)")

    def fetch_oi(self, retries=4):
        if self.mock:
            return _mock_data()

        last_err = None
        for attempt in range(1, retries + 1):
            try:
                r = self.session.get(OI_URL, timeout=20)
                if r.status_code == 200:
                    payload = r.json()
                    rows = payload.get("data", payload) or []
                    if rows:
                        return rows
                    last_err = "empty data"
                else:
                    last_err = f"HTTP {r.status_code}"
            except Exception as e:  # noqa: BLE001
                last_err = repr(e)

            log.warning("Fetch fail (try %d/%d): %s — session refresh...",
                        attempt, retries, last_err)
            time.sleep(3 * attempt)
            try:
                self._new_session()
            except Exception as e:  # noqa: BLE001
                log.warning("Session refresh fail: %r", e)

        raise RuntimeError(f"NSE se data nahi mila: {last_err}")


_MOCK_SYMS = ["RELIANCE", "TCS", "TATASTEEL", "HDFCBANK", "INFY",
              "SBIN", "ICICIBANK", "AXISBANK", "ITC", "WIPRO"]
_MOCK_PREV = {s: random.randint(500_000, 5_000_000) for s in _MOCK_SYMS}
_MOCK_CUR = dict(_MOCK_PREV)


def _mock_data():
    """Kal ka closing OI fix rehta hai, aaj ka OI dheere-dheere move karta hai."""
    out = []
    for s in _MOCK_SYMS:
        _MOCK_CUR[s] = int(_MOCK_CUR[s] * random.uniform(0.995, 1.008))
        out.append({
            "symbol": s,
            "latestOI": _MOCK_CUR[s],
            "prevOI": _MOCK_PREV[s],
            "volume": random.randint(1000, 90000),
            "valueInCr": round(random.uniform(10, 900), 2),
            "underlyValue": round(random.uniform(100, 4000), 2),
        })
    return out


# ------------------------------------------------------------ DATA SHAPING --


def baseline_file(d: date):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, f"baseline_{d.strftime('%Y-%m-%d')}.json")


def load_baseline(d: date):
    p = baseline_file(d)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_baseline(d: date, baseline: dict):
    with open(baseline_file(d), "w", encoding="utf-8") as f:
        json.dump(baseline, f, indent=1)


def to_dataframe(rows, baseline=None, d=None):
    """
    baseline = kal ka closing OI, har symbol ke liye FIX.
    9:15 par pehli baar lock hota hai, phir poore din wahi denominator use hota hai.

        pctChangeOI = (latestOI - baselineOI) / baselineOI * 100
    """
    df = pd.DataFrame(rows)
    if df.empty:
        return df, (baseline or {})

    df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
    df["latestOI"] = pd.to_numeric(df.get("latestOI"), errors="coerce")
    df["prevOI"] = pd.to_numeric(df.get("prevOI"), errors="coerce")

    baseline = dict(baseline or {})

    # Naye symbols ka baseline lock karo (NSE ka prevOI = kal ka closing OI)
    for sym, prev in zip(df["symbol"], df["prevOI"]):
        if sym not in baseline and pd.notna(prev) and prev > 0:
            baseline[sym] = float(prev)

    df["baselineOI"] = df["symbol"].map(baseline)

    base = pd.to_numeric(df["baselineOI"], errors="coerce").replace(0, pd.NA)
    df["changeInOI"] = (df["latestOI"] - base)
    df["pctChangeOI"] = (df["changeInOI"] / base * 100).round(2)

    for c in SNAPSHOT_COLS:
        if c not in df.columns:
            df[c] = pd.NA

    df = df[SNAPSHOT_COLS].copy()
    if d is not None:
        save_baseline(d, baseline)
    return df.sort_values("symbol").reset_index(drop=True), baseline


def slot_labels():
    """['09.15', '09.20', ... '15.30']"""
    out = []
    t = at_ist(today_ist(), MARKET_OPEN)
    end = at_ist(today_ist(), MARKET_CLOSE)
    while t <= end:
        out.append(t.strftime("%H.%M"))
        t += timedelta(minutes=INTERVAL_MIN)
    return out


def excel_path_for(d: date):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, f"NSE_OI_{d.strftime('%Y-%m-%d')}.xlsx")


# ---------------------------------------------------------------- EXCEL IO --


def write_snapshot(path, sheet_name, df):
    mode = "a" if os.path.exists(path) else "w"
    kw = {"if_sheet_exists": "replace"} if mode == "a" else {}
    with pd.ExcelWriter(path, engine="openpyxl", mode=mode, **kw) as xw:
        df.to_excel(xw, sheet_name=sheet_name, index=False)
    _style_sheet(path, sheet_name)
    log.info("Sheet '%s' saved — %d rows -> %s", sheet_name, len(df),
             os.path.basename(path))


def _style_sheet(path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = str(col[0].value or "")
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(header) + 4)
    ws.freeze_panes = "B2"
    wb.save(path)


def build_conclusion(path, change_mode=CHANGE_MODE):
    """Saari time-sheets padho aur ek CONCLUSION sheet banao."""
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    wb = load_workbook(path, read_only=True)
    time_sheets = [s for s in wb.sheetnames if _is_time_sheet(s)]
    wb.close()
    if not time_sheets:
        log.warning("Koi time sheet nahi mili — CONCLUSION skip")
        return None

    time_sheets.sort(key=lambda s: s.replace(".", ""))

    pct_frames, oi_frames = {}, {}
    for sh in time_sheets:
        d = pd.read_excel(path, sheet_name=sh)
        if d.empty or "symbol" not in d.columns:
            continue
        d = d.drop_duplicates(subset="symbol").set_index("symbol")
        pct_frames[sh] = pd.to_numeric(d.get("pctChangeOI"), errors="coerce")
        oi_frames[sh] = pd.to_numeric(d.get("latestOI"), errors="coerce")

    if not pct_frames:
        return None

    if change_mode == "interval":
        oi = pd.DataFrame(oi_frames)
        table = (oi.pct_change(axis=1) * 100).round(2)
        table.iloc[:, 0] = pd.DataFrame(pct_frames).iloc[:, 0]  # pehla slot day-change
    else:
        table = pd.DataFrame(pct_frames).round(2)

    table = table.sort_index()
    table.index.name = "SYMBOL"
    table = table.reset_index()

    with pd.ExcelWriter(path, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as xw:
        table.to_excel(xw, sheet_name="CONCLUSION", index=False)

    _style_conclusion(path, len(time_sheets))
    log.info("CONCLUSION ready — %d symbols x %d slots", len(table), len(time_sheets))
    return path


def _is_time_sheet(name):
    parts = name.split(".")
    return (len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit()
            and len(parts[0]) == 2 and len(parts[1]) == 2)


def _style_conclusion(path, n_slots):
    wb = load_workbook(path)
    ws = wb["CONCLUSION"]

    head_fill = PatternFill("solid", fgColor="C00000")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green = PatternFill("solid", fgColor="E2EFDA")
    red = PatternFill("solid", fgColor="FCE4E4")

    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(name="Arial", bold=True, size=10)
        for cell in row[1:]:
            cell.font = Font(name="Arial", size=10)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, (int, float)):
                cell.fill = green if cell.value >= 0 else red

    ws.column_dimensions["A"].width = 18
    for i in range(2, n_slots + 2):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 85
    wb.save(path)


def upload_to_s3(path):
    """Optional — sirf tab chalega jab OI_S3_BUCKET env var set ho."""
    if not UPLOAD_TO_S3 or not path or not os.path.exists(path):
        return
    try:
        import boto3
        key = f"{S3_PREFIX.strip('/')}/{os.path.basename(path)}"
        boto3.client("s3", region_name=S3_REGION).upload_file(path, S3_BUCKET, key)
        log.info("S3 upload done -> s3://%s/%s", S3_BUCKET, key)
    except Exception as e:  # noqa: BLE001
        log.warning("S3 upload fail: %r", e)


# ---------------------------------------------------------------- RUNTIME --


def is_trading_day(d: date):
    return d.weekday() < 5 and d.strftime("%Y-%m-%d") not in HOLIDAYS_2026


def run_session(mock=False, change_mode=CHANGE_MODE, until=None, build_concl=True):
    today = today_ist()
    if not mock and not is_trading_day(today):
        log.info("Aaj market band hai (%s) — skip", today)
        return None

    path = excel_path_for(today)
    client = NSEClient(mock=mock)
    slots = slot_labels()
    if until:
        slots = [s for s in slots if s.replace(".", "") <= until.replace(":", "")]
    baseline = load_baseline(today)  # resume ke liye — script restart ho to baseline wahi rahe
    log.info("Session start — %d slots, file: %s", len(slots), path)
    if baseline:
        log.info("Purana baseline mila — %d symbols locked", len(baseline))

    for label in slots:
        hh, mm = label.split(".")
        target = at_ist(today, dtime(int(hh), int(mm)))

        if not mock:
            wait = (target - now_ist()).total_seconds()
            if wait > 0:
                log.info("Wait for %s (%d sec)...", label, int(wait))
                time.sleep(wait)
            elif wait < -120:
                log.info("Slot %s nikal gaya — skip", label)
                continue
            time.sleep(3)  # NSE ko thoda time do data update karne ka

        try:
            df, baseline = to_dataframe(client.fetch_oi(), baseline, today)
            if df.empty:
                log.warning("%s: khaali data", label)
                continue
            write_snapshot(path, label, df)
            upload_to_s3(path)
        except Exception as e:  # noqa: BLE001
            log.error("%s: FAIL — %r", label, e)

    if build_concl:
        build_conclusion(path, change_mode)
        upload_to_s3(path)
        log.info("Din khatam. File ready: %s", path)
    else:
        log.info("Part done (till %s). File: %s", until, path)
    return path


def run_daemon(change_mode=CHANGE_MODE):
    log.info("Daemon mode ON — roz apne aap chalega. Band karne ke liye Ctrl+C")
    while True:
        now = now_ist()
        if is_trading_day(now.date()) and now.time() < MARKET_CLOSE:
            run_session(change_mode=change_mode)
        nxt = at_ist(now.date() + timedelta(days=1), dtime(9, 10))
        log.info("Agla run: %s", nxt)
        time.sleep(max(60, (nxt - now_ist()).total_seconds()))


def main():
    ap = argparse.ArgumentParser(description="NSE 5-min OI tracker")
    ap.add_argument("--daemon", action="store_true", help="roz apne aap chalao")
    ap.add_argument("--once", action="store_true", help="ek snapshot lo aur ruk jao")
    ap.add_argument("--mock", action="store_true", help="fake data se test karo")
    ap.add_argument("--rebuild", metavar="FILE", help="purani file se CONCLUSION banao")
    ap.add_argument("--until", metavar="HH:MM",
                    help="is time tak hi chalao, phir ruk jao (jaise 12:20)")
    ap.add_argument("--no-conclusion", action="store_true",
                    help="CONCLUSION sheet mat banao (jab din adhoora ho)")
    ap.add_argument("--mode", choices=["day", "interval"], default=CHANGE_MODE,
                    help="day = vs prev day OI | interval = vs pichhle 5 min")
    args = ap.parse_args()

    setup_logging()

    if args.rebuild:
        build_conclusion(args.rebuild, args.mode)
        return

    if args.once:
        client = NSEClient(mock=args.mock)
        today = today_ist()
        df, _ = to_dataframe(client.fetch_oi(), load_baseline(today), today)
        label = now_ist().strftime("%H.%M")
        path = excel_path_for(today)
        write_snapshot(path, label, df)
        print(df.head(15).to_string(index=False))
        return

    if args.daemon:
        run_daemon(args.mode)
    else:
        run_session(mock=args.mock, change_mode=args.mode,
                    until=args.until, build_concl=not args.no_conclusion)


if __name__ == "__main__":
    main()
